import random
import time
from concurrent.futures.thread import ThreadPoolExecutor

import requests
import json

from fake_useragent import UserAgent
from openpyxl import load_workbook
from requests.exceptions import ProxyError

PROXY_KEY = ""
PROXIES = []
BAD_PROXY = []
USER_AGENT = [
    "Mozilla/5.0 (iPhone; CPU iPhone OS 10_3_1 like Mac OS X) AppleWebKit/603.1.30 (KHTML, like Gecko) Version/10.0 Mobile/14E304 Safari/602.1",
    "Mozilla/5.0 (Linux; U; Android 4.4.2; en-us; SCH-I535 Build/KOT49H) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30",
    "Mozilla/5.0 (Linux; Android 7.0; SM-G930V Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.125 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 7.0; SM-A310F Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.91 Mobile Safari/537.36 OPR/42.7.2246.114996",
    "Opera/9.80 (Android 4.1.2; Linux; Opera Mobi/ADR-1305251841) Presto/2.11.355 Version/12.10",
    "Opera/9.80 (J2ME/MIDP; Opera Mini/5.1.21214/28.2725; U; ru) Presto/2.8.119 Version/11.10",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 7_1_2 like Mac OS X) AppleWebKit/537.51.2 (KHTML, like Gecko) OPiOS/10.2.0.93022 Mobile/11D257 Safari/9537.53",
    "Mozilla/5.0 (Android 7.0; Mobile; rv:54.0) Gecko/54.0 Firefox/54.0",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 10_3_2 like Mac OS X) AppleWebKit/603.2.4 (KHTML, like Gecko) FxiOS/7.5b3349 Mobile/14F89 Safari/603.2.4",
    "Mozilla/5.0 (Linux; U; Android 7.0; en-US; SM-G935F Build/NRD90M) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 UCBrowser/11.3.8.976 U3/0.8.0 Mobile Safari/534.30",
    "Mozilla/5.0 (Linux; Android 6.0.1; SM-G920V Build/MMB29K) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.98 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 5.1.1; SM-N750K Build/LMY47X; ko-kr) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Mobile Safari/537.36 Puffin/6.0.8.15804AP",
    "Mozilla/5.0 (Linux; Android 7.0; SAMSUNG SM-G955U Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) SamsungBrowser/5.4 Chrome/51.0.2704.106 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 6.0; Lenovo K50a40 Build/MRA58K) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.137 YaBrowser/17.4.1.352.00 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; U; Android 7.0; en-us; MI 5 Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/53.0.2785.146 Mobile Safari/537.36 XiaoMi/MiuiBrowser/9.0.3",
    "Mozilla/5.0 (compatible; MSIE 10.0; Windows Phone 8.0; Trident/6.0; IEMobile/10.0; ARM; Touch; Microsoft; Lumia 950)",
    "Mozilla/5.0 (Windows Phone 10.0; Android 6.0.1; Microsoft; Lumia 950) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Mobile Safari/537.36 Edge/15.14977",
    "Mozilla/5.0 (BB10; Kbd) AppleWebKit/537.35+ (KHTML, like Gecko) Version/10.3.3.2205 Mobile Safari/537.35+"

    ]


def get_proxy(proxies):
    if len(proxies) == 0:
        proxies += get_proxies()
    proxy = proxies.pop()
    if proxy['ip'] not in BAD_PROXY:
        BAD_PROXY.append(proxy['ip'])
        return proxy
    else:
        return get_proxy(proxies)


def get_data_visit(url, attempt=0, proxy=None):
    ua = UserAgent()

    if attempt > 25:
        return None, None
    if proxy is None:
        proxy = get_proxy(PROXIES)

    proxy_str = f"{proxy['ip']}:{proxy['port']}"
    if proxy['socks4']:
        proxies = {'https': f'socks4://{proxy_str}', 'socks4': f'socks4://{proxy_str}'}
    elif proxy['http']:
        proxies = {'https': f'http://{proxy_str}', 'http': f'socks4://{proxy_str}'}
    elif proxy['socks5']:
        proxies = {'https': f'socks5://{proxy_str}', 'http': f'socks5://{proxy_str}'}
    else:
        proxies = {'https': f'https://{proxy_str}', 'http': f'https://{proxy_str}'}
    try:
        # user_agent = ua.random
        user_agent = random.choice(USER_AGENT)
        print(user_agent)
        headers = {"content-type": "application/json",
                   "upgrade-insecure-requests": "1",
                   "sec-fetch-user": "?1",
                   "sec-fetch-site": None,
                   "sec-fetch-mode": "navigate",
                   "sec-fetch-dest": "document",
                   "dnt": "1",
                   "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
                   "accept-language": "ru",
                   "accept-encoding": "gzip, deflate, br",
                   "cache-control": "max-age=0",
                    # "cookie": 'locale=ru-ru; .AspNetCore.Antiforgery.xd9Q-ZnrZJo=CfDJ8MZK45L5wIRGm3Bn15aGrTzfAK0wMX4FsTJ2jBxDtNhmrJa92S-8dVbXyZL-nOvrNM7MJeGXMVKvD5oDD_hNduCMz8NtjjPfdpkUwq4jiEyjjOkK-3iHhIqXt9kQNiz2SGHqVDbL77KyUrnBA6IH_pI; sgID=957eea4a-cc39-0a22-f941-f3ec4aae8ed9; loyal-user={%22date%22:%222021-07-27T06:50:49.944Z%22%2C%22isLoyal%22:false}; _vwo_uuid_v2=D837318A318B5FCF31A5B089D14BC55CF|4e74bff94932946dae39dd6d33f6adcc; _ga=GA1.2.1083985613.1627368650; _gid=GA1.2.2002213526.1627368650; fsrndid=false; _gcl_au=1.1.488091802.1627368651; sc_is_visitor_unique=rx8617147.1627368651.11FCCE4E982D4FCF3418BF6F9DF51405.1.1.1.1.1.1.1.1.1; _vis_opt_s=1%7C; _vis_opt_test_cookie=1; _vwo_uuid=D837318A318B5FCF31A5B089D14BC55CF; _vwo_ds=3%241627368648%3A46.09853704%3A%3A; _vwo_sn=0%3A1; _gat=1; _wingify_pc_uuid=43f2464b467c49e4a224f565428cadf8; pxcts=fbf7e080-eea6-11eb-85e2-752dca1d0629; _pxvid=fbf72068-eea6-11eb-928f-0242ac120011; _fbp=fb.1.1627368653440.181636928; _hjid=aa4bec50-dd4c-4a0c-8b99-64cc2c89404c; _hjFirstSeen=1; __qca=P0-833477457-1627368653547; _px3=e5a493258d77f77c86c58dd50b94024046cbfd2d86174282a9d83de854d48f98:gIT6NI+XXTfFz4OFeKc0t0GisgMDVDnF5LM2srS67r5qIWpjf5fSJmzUONDBEGF+OHh/nPGfuXxn+r3UhB/RIQ==:1000:KI4a7pMMqETtm0vSNE/4zaMGZNdA31U0+VgA6oQjxvTGCU76ekuXYxEUVGUBQZSxKqXn+tKzkuN8bFHhLG+WdY71adugVt9DesmsLgsaeKREN2JSerujkcJnroko2t2Buqw///VC4VS6QnRwpPMP+LL0y5eLBK33J2pUxJnsuQck7IygmlwrVy3srkWZImkLr08pVbGwT3VWGKOHITuqHg==; wingify_donot_track_actions=0; _pk_id.1.fd33=dc1b02672eb88af3.1627368651.1.1627368655.1627368651.; _pk_ses.1.fd33=*; SNS=1; _sn_m={"r":{"n":1},"gi":{"countryCode":"RU","country":"Russia","lt":59.8944,"lg":30.2642,"gi":[6255148,2017370,498817,536203]}}; _sn_n={"a":{"i":"1e137c71-fe1c-4fb4-90b1-4a2aae5dc9b3"}}; _sn_a={"a":{"s":1627368655829},"v":"b2f4167f-bb52-4333-b8ba-e927471637fb"}; visitor_id597341=722031304; visitor_id597341-hash=01a922f673fa80488f752b78250892f3beed4f573b1b829d45e54f575d25f27a2a107c192bfd35d2eab76bedd0cea785bc057b7b; intercom-id-e74067abd037cecbecb0662854f02aee12139f95=98c7f16c-81bf-4c6f-8494-29fb6ae9d62a; intercom-session-e74067abd037cecbecb0662854f02aee12139f95=',
                   "user-agent": user_agent}
        print(proxies)
        response = requests.get(url, headers=headers, proxies=proxies, timeout=15)

        if response.status_code == 403:
            print(403)
            return get_data_visit(url, attempt + 1)
        if response.status_code == 500:
            return {'1': "Not found", '2': "Not found", '3': "Not found", '4': "Not found", '5': "Not found",
                    '6': "Not found"}, proxy
        if response.ok:
            response_text = response.text
            try:
                data_visit_from = response_text.find("Sw.preloadedData")
                data_visit_to = response_text.find("Sw.period")
                json_text = json.loads(response_text[data_visit_from + 18:data_visit_to - 10])
                weekly_traffic_numbers = json_text['overview']['EngagementsSimilarweb']['WeeklyTrafficNumbers']
            except Exception:
                weekly_traffic_numbers = None
            if weekly_traffic_numbers is None or len(weekly_traffic_numbers) == 0:
                return {'1': "-", '2': "-", '3': "-", '4': "-", '5': "-", '6': "-"}, proxy
            return weekly_traffic_numbers, proxy
    except ProxyError as e:
        print(e)
        return get_data_visit(url, attempt + 1)

    except Exception as e:
        print(e)
        return get_data_visit(url, attempt + 1)


def get_proxies():
    try:
        new_proxy = requests.get(
            "https://api.best-proxies.ru/proxylist.json?key=%s&limit=5&type=http,https" % PROXY_KEY,
            timeout=60)
        return new_proxy.json()
    except Exception:
        time.sleep(1)
        return get_proxies()


def read_and_write():
    wb = load_workbook(filename='smi_owners.xlsx')
    sheet = wb['smi_owners']
    ws = wb.active
    return wb, sheet, ws


if __name__ == '__main__':
    wb = load_workbook(filename='smi_owners.xlsx')
    sheet = wb['smi_owners']
    ws = wb.active

    l = ["D", "E", "F", "G", "H", "I", "J"]
    is_add_date = False
    # proxy = get_proxy(PROXIES)
    proxy = None
    result = {}
    pool_source = ThreadPoolExecutor(6)

    # for row in range(2, 4):
    for row in range(2, sheet.max_row - 1):
        if sheet["D%s" % row].value is not None:
            continue
        print(row)
        website = sheet["B%s" % row].value
        url = "https://www.similarweb.com/ru/website/" + website.replace("https://", "").replace("http://", "")
        weekly_traffic_numbers, proxy = get_data_visit(url, 0, proxy)
        if weekly_traffic_numbers is not None:
            if not is_add_date:
                j = 0
                for key in weekly_traffic_numbers:
                    ws[l[j] + str(1)] = key
                    j += 1
                is_add_date = True
            i = 0
            for value in weekly_traffic_numbers.values():
                ws[l[i] + str(row)] = value
                i += 1
            wb.save("smi_owners.xlsx")
    wb.save("smi_owners.xlsx")

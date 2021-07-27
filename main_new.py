import random
from concurrent.futures.thread import ThreadPoolExecutor

import requests
import json
from fake_useragent import UserAgent

from openpyxl import load_workbook
from requests.exceptions import ProxyError

PROXY_KEY = ""
PROXIES = []

USER_AGENT = ["Mozilla/5.0 (iPhone; CPU iPhone OS 10_3_1 like Mac OS X) AppleWebKit/603.1.30 (KHTML, like Gecko) Version/10.0 Mobile/14E304 Safari/602.1",
             "Mozilla/5.0 (Linux; U; Android 4.4.2; en-us; SCH-I535 Build/KOT49H) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30",
             "Mozilla/5.0 (Linux; Android 7.0; SM-G930V Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.125 Mobile Safari/537.36",
             "Mozilla/5.0 (Linux; Android 7.0; SM-A310F Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.91 Mobile Safari/537.36 OPR/42.7.2246.114996",
             "Opera/9.80 (Android 4.1.2; Linux; Opera Mobi/ADR-1305251841) Presto/2.11.355 Version/12.10",
             "Opera/9.80 (J2ME/MIDP; Opera Mini/5.1.21214/28.2725; U; ru) Presto/2.8.119 Version/11.10",
             "Mozilla/5.0 (iPhone; CPU iPhone OS 7_1_2 like Mac OS X) AppleWebKit/537.51.2 (KHTML, like Gecko) OPiOS/10.2.0.93022 Mobile/11D257 Safari/9537.53",
             "Mozilla/5.0 (Android 7.0; Mobile; rv:54.0) Gecko/54.0 Firefox/54.0",
             "Mozilla/5.0 (iPhone; CPU iPhone OS 10_3_2 like Mac OS X) AppleWebKit/603.2.4 (KHTML, like Gecko) FxiOS/7.5b3349 Mobile/14F89 Safari/603.2.4",
             "Mozilla/5.0 (Linux; U; Android 7.0; en-US; SM-G935F Build/NRD90M) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 UCBrowser/11.3.8.976 U3/0.8.0 Mobile Safari/534.30",
             "Mozilla/5.0 (Linux; Android 6.0.1; SM-G920V Build/MMB29K) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.98 Mobile Safari/537.36",
             "Mozilla/5.0 (Linux; Android 5.1.1; SM-N750K Build/LMY47X; ko-kr) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Mobile Safari/537.36 Puffin/6.0.8.15804AP",
             "Mozilla/5.0 (Linux; Android 7.0; SAMSUNG SM-G955U Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) SamsungBrowser/5.4 Chrome/51.0.2704.106 Mobile Safari/537.36",
             "Mozilla/5.0 (Linux; Android 6.0; Lenovo K50a40 Build/MRA58K) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.137 YaBrowser/17.4.1.352.00 Mobile Safari/537.36",
             "Mozilla/5.0 (Linux; U; Android 7.0; en-us; MI 5 Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/53.0.2785.146 Mobile Safari/537.36 XiaoMi/MiuiBrowser/9.0.3",
             "Mozilla/5.0 (compatible; MSIE 10.0; Windows Phone 8.0; Trident/6.0; IEMobile/10.0; ARM; Touch; Microsoft; Lumia 950)",
             "Mozilla/5.0 (Windows Phone 10.0; Android 6.0.1; Microsoft; Lumia 950) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Mobile Safari/537.36 Edge/15.14977",
             "Mozilla/5.0 (BB10; Kbd) AppleWebKit/537.35+ (KHTML, like Gecko) Version/10.3.3.2205 Mobile Safari/537.35+"

             ]

def get_proxy(proxies):
    if len(proxies) == 0:
        proxies += get_proxies()
    return proxies.pop()


def get_data_visit(row, result, url, attempt=0, proxy=None):
    ua = UserAgent()

    if attempt > 25:
        return None, None
    if proxy is None:
        proxy = get_proxy(PROXIES)

    proxy_str = f"{proxy['ip']}:{proxy['port']}"
    if proxy['socks4']:
        proxies = {'https': f'socks4://{proxy_str}', 'socks4': f'socks4://{proxy_str}'}
    elif proxy['http']:
        proxies = {'https': f'http://{proxy_str}', 'http': f'socks4://{proxy_str}'}
    elif proxy['socks5']:
        proxies = {'https': f'socks5://{proxy_str}', 'http': f'socks5://{proxy_str}'}
    else:
        proxies = {'https': f'https://{proxy_str}', 'http': f'https://{proxy_str}'}
    try:
        # user_agent = ua.random
        user_agent = random.choice(USER_AGENT)
        print(user_agent)
        headers = {"content-type": "application/json",
                   "upgrade-insecure-requests": "1",
                   "sec-fetch-user": "?1",
                   "sec-fetch-site": None,
                   "sec-fetch-mode": "navigate",
                   "sec-fetch-dest": "document",
                   "dnt": "1",
                   "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
                   "accept-language": "ru",
                    "accept-encoding": "gzip, deflate, br",
                   "cache-control": "max-age=0",
                   "user-agent":user_agent}
        print(proxies)
        response = requests.get(url, headers=headers, proxies=proxies, timeout=10)

        if response.status_code == 403:
            print(403)
            return get_data_visit(row, result, url, attempt+1)
        if response.status_code == 500:
            result.update({row: {'1': "Not found", '2': "Not found", '3': "Not found", '4': "Not found", '5': "Not found", '6': "Not found"}})
            return

        if response.ok:
            response_text = response.text
            try:
                data_visit_from = response_text.find("Sw.preloadedData")
                data_visit_to = response_text.find("Sw.period")
                json_text = json.loads(response_text[data_visit_from + 18:data_visit_to - 10])
                weekly_traffic_numbers = json_text['overview']['EngagementsSimilarweb']['WeeklyTrafficNumbers']
            except Exception:
                weekly_traffic_numbers = None
            if weekly_traffic_numbers is None or len(weekly_traffic_numbers) == 0:
                weekly_traffic_numbers=  {'1': "-", '2': "-", '3': "-", '4': "-", '5': "-", '6': "-"}
            result.update({row: weekly_traffic_numbers})
            return
    except ProxyError as e:
        print(e)
        return get_data_visit(row, result, url, attempt+1)

    except Exception as e:
        print(e)
        return get_data_visit(row, result, url, attempt+1)


def get_proxies():
    new_proxy = requests.get(
        "https://api.best-proxies.ru/proxylist.json?key=%s&speed=1&limit=100" % PROXY_KEY,
        timeout=60)
    return new_proxy.json()


def read_and_write():
    wb = load_workbook(filename='smi_owners.xlsx')
    sheet = wb['smi_owners']
    ws = wb.active
    return wb, sheet, ws


if __name__ == '__main__':

    wb, sheet, ws = read_and_write()

    l = ["D", "E", "F", "G", "H", "I", "J"]
    is_add_date = False
    # proxy = get_proxy(PROXIES)
    proxy = None
    result = {}
    pool_source = ThreadPoolExecutor(35)

    # for row in range(2, 4):
    result = {}
    try:
        for row in reversed(range(2, sheet.max_row - 1)):
            if sheet["D%s" % row].value is not None:
                continue
            print(row)
            website = sheet["B%s" % row].value
            url = "https://www.similarweb.com/ru/website/" + website.replace("https://", "").replace("http://", "")
            pool_source.submit(get_data_visit, row, result, url, 0, proxy)
            if row % 50 == 0:
                count = 0
                pool_source.shutdown()
                try:
                    dict_result = result.copy()
                    for key, value in dict_result.items():
                        if value is not None and len(value) > 0:
                            i = 0
                            count += 1
                            for v in value.values():
                                ws[l[i] + str(key)] = v
                                i += 1
                    try:
                        print("SAVE: " + str(count))
                        wb.save("smi_owners.xlsx")
                    except Exception as e:
                        print("CANNOTSAVE " + str(e))
                except Exception as e:
                    print("CAN NOT SAVE" + str(e))
                result = {}
                pool_source = ThreadPoolExecutor(10)
    except Exception:
        pass
    wb, sheet, ws = read_and_write()

    l = ["D", "E", "F", "G", "H", "I", "J"]
    # proxy = get_proxy(PROXIES)
    proxy = None
    result = {}
    pool_source = ThreadPoolExecutor(5)

    try:
        result = {}
        for row in range(2, sheet.max_row - 1):
            if sheet["D%s" % row].value is not None:
                continue
            print(row)
            website = sheet["B%s" % row].value
            url = "https://www.similarweb.com/ru/website/" + website.replace("https://", "").replace("http://", "")
            pool_source.submit(get_data_visit, row, result, url, 0, proxy)
            if row % 5 == 0:
                count = 0
                pool_source.shutdown()
                try:
                    dict_result = result.copy()
                    for key, value in dict_result.items():
                        if value is not None and len(value) > 0:
                            i = 0
                            count += 1
                            for v in value.values():
                                ws[l[i] + str(key)] = v
                                i += 1
                    try:
                        print("SAVE: " + str(count))
                        wb.save("smi_owners.xlsx")
                    except Exception as e:
                        print("CANNOTSAVE " + str(e))
                except Exception as e:
                    print("CAN NOT SAVE" + str(e))
                result = {}
                pool_source = ThreadPoolExecutor(10)
    except Exception as e:
        pass

    wb, sheet, ws = read_and_write()
    proxy = None
    result = {}
    pool_source = ThreadPoolExecutor(5)

    try:
        result = {}
        for row in range(2, sheet.max_row - 1):
            if sheet["D%s" % row].value is not None:
                continue
            print(row)
            website = sheet["B%s" % row].value
            url = "https://www.similarweb.com/ru/website/" + website.replace("https://", "").replace("http://", "")
            pool_source.submit(get_data_visit, row, result, url, 0, proxy)
            if row % 5 == 0:
                count = 0
                pool_source.shutdown()
                try:
                    dict_result = result.copy()
                    for key, value in dict_result.items():
                        if value is not None and len(value) > 0:
                            i = 0
                            count += 1
                            for v in value.values():
                                ws[l[i] + str(key)] = v
                                i += 1
                    try:
                        print("SAVE: " + str(count))
                        wb.save("smi_owners.xlsx")
                    except Exception as e:
                        print("CANNOTSAVE " + str(e))
                except Exception as e:
                    print("CAN NOT SAVE" + str(e))
                result = {}
                pool_source = ThreadPoolExecutor(10)
    except Exception as e:
        pass


    wb, sheet, ws = read_and_write()
    proxy = None
    result = {}
    pool_source = ThreadPoolExecutor(15)

    try:
        result = {}
        for row in range(2, sheet.max_row - 1):
            if sheet["D%s" % row].value is not None:
                continue
            print(row)
            website = sheet["B%s" % row].value
            url = "https://www.similarweb.com/ru/website/" + website.replace("https://", "").replace("http://", "")
            pool_source.submit(get_data_visit, row, result, url, 0, proxy)
            if row % 5 == 0:
                count = 0
                pool_source.shutdown()
                try:
                    dict_result = result.copy()
                    for key, value in dict_result.items():
                        if value is not None and len(value) > 0:
                            i = 0
                            count += 1
                            for v in value.values():
                                ws[l[i] + str(key)] = v
                                i += 1
                    try:
                        print("SAVE: " + str(count))
                        wb.save("smi_owners.xlsx")
                    except Exception as e:
                        print("CANNOTSAVE " + str(e))
                except Exception as e:
                    print("CAN NOT SAVE" + str(e))
                result = {}
                pool_source = ThreadPoolExecutor(10)
    except Exception as e:
        pass





